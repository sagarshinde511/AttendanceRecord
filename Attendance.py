from datetime import datetime, timedelta
import pandas as pd
import streamlit as st
import mysql.connector
from io import BytesIO
import openpyxl
from openpyxl.styles import PatternFill

# Function to apply color in Streamlit table
def highlight_status(val):
    if val == 'Absent':
        return 'background-color: #FF9999'  # Red
    elif val == 'Holiday':
        return 'background-color: #FFFF99'  # Yellow
    return ''
# MySQL connection
conn = mysql.connector.connect(
    host="82.180.143.66",
    user="u263681140_AttendanceInt",
    password="SagarAtten@12345",
    database="u263681140_Attendance"
)
cursor = conn.cursor(dictionary=True)

# Streamlit layout
st.set_page_config(layout="wide")
tabs = st.tabs(["📋 Raw Attendance", "📊 Summary Report"])

# ---------------- TAB 1 ----------------
with tabs[0]:
    query = """
    SELECT ar.RollNo, ar.Date, ar.Time, sd.id, sd.StudentName, sd.Batch
    FROM AttendanceRecord ar
    JOIN Students_Data sd ON ar.RollNo = sd.id
    ORDER BY ar.Date DESC, ar.Time DESC
    """
    cursor.execute(query)
    results = cursor.fetchall()
    df = pd.DataFrame(results)
    st.title("📋 Raw Attendance Log")
    st.dataframe(df)

# ---------------- TAB 2 ----------------
with tabs[1]:
    st.title("📊 Attendance Summary Report")

    # Batch filter
    cursor.execute("SELECT DISTINCT Batch FROM Students_Data")
    batches = [row['Batch'] for row in cursor.fetchall()]
    selected_batch = st.selectbox("Select Batch", batches)

    from_date = st.date_input("From Date", datetime.today() - timedelta(days=7))
    to_date = st.date_input("To Date", datetime.today())

    # Get students in batch
    cursor.execute("SELECT id as RollNo, StudentName FROM Students_Data WHERE Batch = %s", (selected_batch,))
    students = pd.DataFrame(cursor.fetchall())
    students['RollNo'] = students['RollNo'].astype(str)

    # Get date range
    date_range = pd.date_range(start=from_date, end=to_date)
    date_df = pd.DataFrame({'Date': date_range.strftime('%Y-%m-%d')})

    # Full grid
    full_grid = students.assign(key=1).merge(date_df.assign(key=1), on='key').drop('key', axis=1)

    # Get attendance records
    cursor.execute("""
        SELECT RollNo, Date FROM AttendanceRecord
        WHERE Date BETWEEN %s AND %s AND RollNo IN (
            SELECT id FROM Students_Data WHERE Batch = %s
        )
    """, (from_date, to_date, selected_batch))
    attendance = pd.DataFrame(cursor.fetchall())
    attendance['RollNo'] = attendance['RollNo'].astype(str)
    attendance['Date'] = pd.to_datetime(attendance['Date']).dt.strftime('%Y-%m-%d')
    attendance['Status'] = 'Present'

    # Merge
    full_grid['Date'] = pd.to_datetime(full_grid['Date']).dt.strftime('%Y-%m-%d')
    merged = full_grid.merge(attendance, on=['RollNo', 'Date'], how='left')
    merged['Status'] = merged['Status'].fillna('Absent')

    # Pivot
    pivot_df = merged.pivot(index=['RollNo', 'StudentName'], columns='Date', values='Status').reset_index()

    # Mark Holidays
    for col in pivot_df.columns[2:]:
        if all(pivot_df[col] == 'Absent'):
            pivot_df[col] = 'Holiday'

    # Reorder
    pivot_df = pivot_df[['RollNo', 'StudentName'] + sorted(pivot_df.columns[2:])]
    # Sort by RollNo ascending
    pivot_df = pivot_df.sort_values(by="RollNo")


    # Display with colors
    styled_df = pivot_df.style.applymap(highlight_status, subset=pivot_df.columns[2:])
    st.dataframe(styled_df, use_container_width=True)

    # Excel generation with color
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attendance Report"

    for col_num, column_title in enumerate(pivot_df.columns, 1):
        ws.cell(row=1, column=col_num).value = column_title

    for row_idx, row in pivot_df.iterrows():
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx + 2, column=col_idx, value=value)
            if value == 'Absent':
                cell.fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
            elif value == 'Holiday':
                cell.fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")

    # Save to BytesIO
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)

    st.download_button(
        label="📥 Download Excel with Colors",
        data=excel_file,
        file_name=f"Attendance_{selected_batch}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
